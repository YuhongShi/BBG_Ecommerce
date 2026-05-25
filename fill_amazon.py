import sys
import os
import glob

sys.stdout.reconfigure(encoding='utf-8')

FOLDER = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Amazon\2026年4月"
XLSX   = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_April.xlsx"

AMAZON_HEADERS = [
    'Nr', 'Marktplatz', 'Settlement-Periode',
    'Brutto_Sale', 'Erstattung', 'FBA_Gebühr',
    'Commission', 'Commission_rev', 'Promotion',
    'Werbungskosten', 'Lagergebühr', 'USt_Einbeh',
    'Sonstiges', 'Summe'
]

def parse_eur(s):
    if not s or s.strip() == '':
        return 0.0
    s = s.strip()
    if ',' in s:
        # European format: . = thousands sep, , = decimal sep
        s = s.replace('.', '').replace(',', '.')
    # else: standard format (. = decimal sep), leave as-is
    try:
        return float(s)
    except ValueError:
        return 0.0

def fmt(s):
    return s.strip() if s else ''

# ── Step 1: scan txt files ────────────────────────────────────
positive_files = []
skipped_files  = []

txt_files = glob.glob(os.path.join(FOLDER, "*.txt"))
for fpath in txt_files:
    fname = os.path.basename(fpath)
    with open(fpath, encoding='utf-8') as f:
        lines = f.readlines()
    if len(lines) < 2:
        skipped_files.append((fname, 'no data rows'))
        continue
    cols = lines[1].split('\t')
    if len(cols) < 6:
        skipped_files.append((fname, 'bad header'))
        continue
    total_amount = parse_eur(cols[4])
    currency     = fmt(cols[5])
    if currency != 'EUR':
        skipped_files.append((fname, f'currency={currency}'))
        continue
    if total_amount <= 0:
        skipped_files.append((fname, f'total-amount={total_amount} EUR (Amazon 扣款)'))
        continue
    positive_files.append((fpath, fname, lines, total_amount))

print(f"已处理结算文件：{len(positive_files)} 个（正数 EUR）")
print(f"跳过文件：{len(skipped_files)} 个")

# ── Step 2: parse each positive settlement ────────────────────
settlements = []

for fpath, fname, lines, summe in positive_files:
    cols2 = lines[1].split('\t')
    start_raw = fmt(cols2[1])
    end_raw   = fmt(cols2[2])

    def extract_ddmm(s):
        parts = s.split(' ')[0].split('.')
        if len(parts) == 3:
            return parts[0] + '.' + parts[1], parts[2]
        return s, ''

    start_dm, _     = extract_ddmm(start_raw)
    end_dm,   end_y = extract_ddmm(end_raw)
    periode = f"{start_dm}~{end_dm}.{end_y}"

    brutto_sale    = 0.0
    erstattung     = 0.0
    fba            = 0.0
    commission     = 0.0
    commission_rev = 0.0
    promotion      = 0.0
    werbung        = 0.0
    lager          = 0.0
    ust            = 0.0
    sonstiges      = 0.0

    for line in lines[2:]:
        c = line.split('\t')
        if len(c) < 15:
            continue
        tx_type  = fmt(c[6])
        amt_type = fmt(c[12])
        amt_desc = fmt(c[13])
        amt_raw  = fmt(c[14])
        if not amt_raw:
            continue
        amt = parse_eur(amt_raw)
        if amt == 0.0:
            continue

        classified = False

        if tx_type == 'Order' and amt_type == 'ItemPrice':
            brutto_sale += amt
            classified = True
        elif tx_type == 'Refund' and amt_type == 'ItemPrice':
            erstattung += amt
            classified = True
        elif amt_desc in ('FBAPerUnitFulfillmentFee', 'ShippingChargeback'):
            fba += amt
            classified = True
        elif amt_desc == 'Commission' and amt < 0:
            commission += amt
            classified = True
        elif amt_desc == 'Commission' and amt > 0:
            commission_rev += amt
            classified = True
        elif amt_type == 'Promotion':
            promotion += amt
            classified = True
        elif tx_type == 'ServiceFee' and amt_desc == 'Cost of Advertising':
            werbung += amt
            classified = True
        elif tx_type == 'other-transaction' and amt_desc == 'Storage Fee':
            lager += amt
            classified = True
        elif amt_type == 'ItemWithheldTax':
            ust += amt
            classified = True

        if not classified:
            sonstiges += amt

    calc_sum = (brutto_sale + erstattung + fba + commission + commission_rev
                + promotion + werbung + lager + ust + sonstiges)
    diff = abs(calc_sum - summe)
    if diff > 0.05:
        print(f"  [警告] {fname}: 计算合计={calc_sum:.2f} vs Summe={summe:.2f}, 差异={diff:.2f}")
        print(f"         （差异通常由 Amazon 发票费用（Vine/Deal 等）引起，按正常流程继续）")

    settlements.append({
        'fname':        fname,
        'periode':      periode,
        'brutto_sale':  round(brutto_sale,    2),
        'erstattung':   round(erstattung,     2),
        'fba':          round(fba,            2),
        'commission':   round(commission,     2),
        'comm_rev':     round(commission_rev, 2),
        'promotion':    round(promotion,      2),
        'werbung':      round(werbung,        2),
        'lager':        round(lager,          2),
        'ust':          round(ust,            2),
        'sonstiges':    round(sonstiges,      2),
        'summe':        round(summe,          2),
    })

# ── openpyxl ──────────────────────────────────────────────────
import openpyxl
wb = openpyxl.load_workbook(XLSX)
sheet1 = wb['Sheet1']

# Create Amazon sheet if missing
if 'Amazon' not in wb.sheetnames:
    amazon = wb.create_sheet('Amazon')
    for col_idx, header in enumerate(AMAZON_HEADERS, start=1):
        amazon.cell(row=1, column=col_idx).value = header
    print("\n已创建 Amazon sheet。")
else:
    amazon = wb['Amazon']

# ── Step 3: find Sheet1 Amazon rows ───────────────────────────
sheet1_amazon_rows = []
for row in sheet1.iter_rows(min_row=2):
    zv     = row[2].value  # col 3: Zahlungsverkehr
    betrag = row[3].value  # col 4: Betrag
    if zv and 'Amazon' in str(zv) and betrag is not None:
        sheet1_amazon_rows.append({
            'row_num': row[0].row,
            'nr':      row[0].value,
            'betrag':  float(betrag),
        })

# ── Step 4: build Amazon sheet row index by Nr ───────────────
amazon_rows_by_nr = {}
amazon_last_data_row = 1
for row in amazon.iter_rows(min_row=2):
    nr_val = row[0].value
    if nr_val is not None:
        try:
            nr_int = int(nr_val)
            amazon_rows_by_nr[nr_int] = row
            amazon_last_data_row = row[0].row
        except (ValueError, TypeError):
            pass

def write_amazon_row(cells, s, nr):
    cells[0].value  = nr
    cells[1].value  = 'Amazon.de'
    cells[2].value  = s['periode']
    cells[3].value  = s['brutto_sale']
    cells[4].value  = s['erstattung']
    cells[5].value  = s['fba']
    cells[6].value  = s['commission']
    cells[7].value  = s['comm_rev']
    cells[8].value  = s['promotion']
    cells[9].value  = s['werbung']
    cells[10].value = s['lager']
    cells[11].value = s['ust']
    cells[12].value = s['sonstiges']
    cells[13].value = s['summe']

# ── Print summary header ──────────────────────────────────────
print("\n结算汇总：")
print(f"{'Nr':>4} | {'Settlement-Periode':<20} | {'Brutto_Sale':>11} | {'Erstattung':>10} | "
      f"{'FBA':>8} | {'Commission':>10} | {'Comm_rev':>8} | {'Promotion':>9} | "
      f"{'Werbung':>8} | {'Lager':>7} | {'USt':>7} | {'Sonst':>10} | {'Summe':>8} | Sheet1核对")
print('-' * 150)

sheet1_updates = []

for s in settlements:
    # match Sheet1 row by Betrag ≈ Summe
    matched = None
    for r in sheet1_amazon_rows:
        if abs(r['betrag'] - s['summe']) <= 0.02:
            matched = r
            break
    if matched is None:
        print(f"  [警告] 找不到 Sheet1 匹配行 (Summe={s['summe']}), 跳过 {s['fname']}")
        continue

    nr = matched['nr']
    try:
        nr = int(nr)
    except (TypeError, ValueError):
        pass

    print(f"{str(nr):>4} | {s['periode']:<20} | {s['brutto_sale']:>11.2f} | {s['erstattung']:>10.2f} | "
          f"{s['fba']:>8.2f} | {s['commission']:>10.2f} | {s['comm_rev']:>8.2f} | {s['promotion']:>9.2f} | "
          f"{s['werbung']:>8.2f} | {s['lager']:>7.2f} | {s['ust']:>7.2f} | {s['sonstiges']:>10.2f} | "
          f"{s['summe']:>8.2f} | ✓")

    # write to Amazon sheet
    if nr in amazon_rows_by_nr:
        write_amazon_row(amazon_rows_by_nr[nr], s, nr)
    else:
        next_row = amazon_last_data_row + 1
        cells = [amazon.cell(row=next_row, column=c) for c in range(1, 15)]
        write_amazon_row(cells, s, nr)
        amazon_last_data_row = next_row

    brutto_umsatz    = round(s['brutto_sale'] + s['erstattung'], 2)
    andere_gebuehren = round(s['fba'] + s['commission'] + s['comm_rev']
                             + s['promotion'] + s['werbung'] + s['lager']
                             + s['ust'] + s['sonstiges'], 2)
    sheet1_updates.append((nr, matched['row_num'], brutto_umsatz, andere_gebuehren))

# ── Step 5: write Sheet1 cols 5 and 6 ────────────────────────
print("\nSheet1 已填入：")
for nr, row_num, bu, ag in sheet1_updates:
    sheet1.cell(row=row_num, column=5).value = bu
    sheet1.cell(row=row_num, column=6).value = ag
    print(f"  Nr {nr}: Brrutto_Umsatz={bu:.2f}, Andere Gebühren={ag:.2f}")

wb.save(XLSX)
print("\n文件已保存。")

if skipped_files:
    print("\n以下结算为 Amazon 扣款（未填入 Sheet1，请手动处理）：")
    for fname, reason in skipped_files:
        print(f"  - {fname}: {reason}")
