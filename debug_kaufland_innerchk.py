#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Debug InnerChk non-zero rows by printing all transactions for those payout groups."""

import sys, re, argparse
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

# ── paste same helpers from fill_kaufland.py ───────────────────────────────
def parse_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try: return float(s)
    except: return 0.0

def r2(v): return round(v, 2)

def extract_ym(sheet_name):
    m = re.search(r'(\d+)年(\d+)月', sheet_name)
    return (int(m.group(1)), int(m.group(2))) if m else None

def read_sheet_rows(ws):
    hdr = next(
        (r for r in range(1, 12) if str(ws.cell(r, 1).value or "").strip() == "booking_date"),
        None
    )
    if hdr is None:
        return []
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        bt = ws.cell(r, 2).value
        if bt is None:
            continue
        rows.append({
            "bt":      str(bt).strip(),
            "amount":  parse_num(ws.cell(r, 5).value),
            "balance": parse_num(ws.cell(r, 6).value),
            "sgross":  parse_num(ws.cell(r, 9).value),
            "fgross":  parse_num(ws.cell(r, 13).value),
            "vat_pct": parse_num(ws.cell(r, 12).value),
        })
    return rows

def get_prev_tail(ws):
    rows = read_sheet_rows(ws)
    if not rows:
        return 0.0, []
    last_payout_idx = None
    for i, row in enumerate(rows):
        if row["bt"] == "Payout":
            last_payout_idx = i
    if last_payout_idx is None:
        opening = rows[0]["balance"] - rows[0]["amount"]
        return opening, rows
    return rows[last_payout_idx]["balance"], rows[last_payout_idx + 1:]

def classify_tx(bt, amt, sg, fg):
    if bt.startswith("Freigabe") or bt.startswith("Release order"):
        return [("brutto", sg), ("commission", -fg)]
    if (bt.startswith("Erstattung") or bt.startswith("Partial Refund")
            or bt.startswith("Storno")):
        return [("erstattung", amt)]
    bt_lo = bt.lower()
    if bt_lo.startswith("fees for cancelled") or bt_lo.startswith("gebühren für stornierte"):
        return [("steuer", amt)]
    canonical = bt
    for prefix in ("Umsatzsteuer ", "Netto ", "Net "):
        if canonical.startswith(prefix):
            canonical = canonical[len(prefix):]
            break
    canonical_lo = canonical.lower()
    if "grundgebühr" in canonical_lo:
        return [("marktplatz", amt)]
    if "gutschein" in canonical_lo or "voucher" in canonical_lo:
        return [("gutschein", amt)]
    return [("sonstiges", amt)]

# ── target Nr rows to investigate ─────────────────────────────────────────
# Nr -> (country, summe_eur) from last run output
TARGET_SUMME = {
    26:  ("法国站",    36.68),
    46:  ("捷克站",   793.79),
    50:  ("波兰站",  1141.61),
    57:  ("意大利站", 131.14),
    58:  ("斯洛伐克站", 237.33),
    60:  ("奥地利站",  689.51),
}

KAUFLAND_FILE = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Kaufland\2026年\5月\kaufland-BBG 26年5月回款分佣及vat应结算金额-41772.49欧元(施总)(已发送确认中).xlsx"
PREV_KL_FILE  = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Kaufland\2026年\4月\kaufland-BBG 26年4月回款分佣及vat应结算金额-30552.68欧元(施总)(已发送确认中).xlsx"
UMSATZ_FILE   = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_Mai.xlsx"

COUNTRY_CFG = {
    "德国站":     {"currency": "EUR", "behalten": True},
    "奥地利站":   {"currency": "EUR", "behalten": False},
    "法国站":     {"currency": "EUR", "behalten": False},
    "斯洛伐克站": {"currency": "EUR", "behalten": False},
    "意大利站":   {"currency": "EUR", "behalten": False},
    "捷克站":     {"currency": "CZK", "behalten": False},
    "波兰站":     {"currency": "PLN", "behalten": False},
}
COUNTRY_S1_LABELS = {
    "德国站":     ["Kaufland", "Kaufland-DE"],
    "奥地利站":   ["Kaufland", "Kaufland-AT"],
    "法国站":     ["Kaufland", "Kaufland-FR"],
    "斯洛伐克站": ["Kaufland"],
    "意大利站":   ["Kaufland"],
    "捷克站":     ["KTO"],
    "波兰站":     ["IBAN", "IBAN(Unbekannt)"],
}
ALL_KL_LABELS = {"Kaufland","Kaufland-DE","Kaufland-AT","Kaufland-FR","KTO","IBAN","IBAN(Unbekannt)"}
FUZZY_TOL = 0.06

wb_kl   = openpyxl.load_workbook(KAUFLAND_FILE, data_only=True)
wb_prev = openpyxl.load_workbook(PREV_KL_FILE,  data_only=True)
wb_us   = openpyxl.load_workbook(UMSATZ_FILE)
ws_s1   = wb_us["Sheet1"]

rates = {"EUR": 1.0, "PLN": 0.2356, "CZK": 0.0411}
ws_sum = wb_kl["合计应结算"]
for row in range(14, 22):
    label = str(ws_sum.cell(row, 1).value or "")
    rtext = str(ws_sum.cell(row, 3).value or "")
    m = re.search(r"\*\s*(\d+\.?\d+)", rtext)
    if m:
        rate = float(m.group(1))
        if "PLN" in label: rates["PLN"] = rate
        if "CZK" in label: rates["CZK"] = rate

# Build Sheet1 lookup
s1_entries = []
for r in range(2, ws_s1.max_row + 1):
    nr     = ws_s1.cell(r, 1).value
    z      = ws_s1.cell(r, 3).value
    betrag = ws_s1.cell(r, 4).value
    if not nr or not z or not betrag:
        continue
    zs = str(z)
    if zs in ALL_KL_LABELS:
        s1_entries.append({"nr": int(nr), "label": zs, "betrag": float(betrag)})

# Build country→sheets map
country_sheet_map = {}
for sn in wb_kl.sheetnames:
    country = next((k for k in COUNTRY_CFG if k in sn), None)
    if not country:
        continue
    ym = extract_ym(sn)
    if ym is None:
        continue
    country_sheet_map.setdefault(country, []).append((ym, sn))
for c in country_sheet_map:
    country_sheet_map[c].sort(key=lambda x: x[0])

# ── Process and find matching groups ──────────────────────────────────────
target_countries = set(v[0] for v in TARGET_SUMME.values())

for country in target_countries:
    cfg      = COUNTRY_CFG[country]
    currency = cfg["currency"]
    rate     = rates[currency]

    sheets_for_country = country_sheet_map.get(country, [])
    if not sheets_for_country:
        continue

    current_sn = sheets_for_country[-1][1]
    ws_cur     = wb_kl[current_sn]

    prev_bal_local = 0.0
    prev_tail_rows = []

    if len(sheets_for_country) >= 2:
        prev_sn = sheets_for_country[-2][1]
        prev_bal_local, prev_tail_rows = get_prev_tail(wb_kl[prev_sn])
    elif country == "德国站":
        pass  # skip DE for this debug

    cur_rows = read_sheet_rows(ws_cur)
    groups   = []
    head     = list(prev_tail_rows)
    for row in cur_rows:
        if row["bt"] == "Payout":
            groups.append({"txns": head, "payout": row})
            head = []
        else:
            head.append(row)

    # Derive tax rate
    tax_rate = 0.19
    for row in prev_tail_rows + cur_rows:
        bt = row["bt"]
        if (bt.startswith("Freigabe") or bt.startswith("Release order")) and row["vat_pct"] > 0:
            tax_rate = row["vat_pct"] / 100
            break

    prev_bal = prev_bal_local

    for g in groups:
        txns         = g["txns"]
        payout       = g["payout"]
        payout_local = abs(payout["amount"])
        payout_eur   = r2(payout_local * rate)
        kto_gut      = r2(prev_bal * rate)
        behalten_val = r2(-payout["balance"] * rate)
        prev_bal     = payout["balance"]

        # Find which target Nr this corresponds to
        matched_nr = None
        for nr, (c, target_summe) in TARGET_SUMME.items():
            if c == country and abs(payout_eur - target_summe) < 0.05:
                # Also check S1 entry for this Nr
                s1 = next((e for e in s1_entries if e["nr"] == nr), None)
                if s1:
                    matched_nr = nr
                    break

        if matched_nr is None:
            continue

        # --- Print detailed breakdown ---
        target_summe = TARGET_SUMME[matched_nr][1]
        print(f"\n{'='*70}")
        print(f"Nr={matched_nr}  {country}  sheet={current_sn}")
        print(f"  payout_local={payout_local:.4f} {currency}  rate={rate}  payout_eur={payout_eur:.2f}")
        print(f"  kto_gut={kto_gut:.2f}  behalten={behalten_val:.2f}  tax_rate={tax_rate}")
        print(f"\n  Transactions in this payout group ({len(txns)} rows):")
        print(f"  {'booking_text':<50} {'amount':>10} {'sgross':>10} {'fgross':>10} -> classified as")

        totals = {"brutto":0.0,"erstattung":0.0,"marktplatz":0.0,
                  "commission":0.0,"gutschein":0.0,"steuer":0.0,"sonstiges":0.0}
        for tx in txns:
            classified = classify_tx(tx["bt"], tx["amount"], tx["sgross"], tx["fgross"])
            for field, val in classified:
                totals[field] += val
            class_str = ", ".join(f"{f}={v:.4f}" for f, v in classified)
            print(f"  {tx['bt']:<50} {tx['amount']:>10.4f} {tx['sgross']:>10.4f} {tx['fgross']:>10.4f} -> {class_str}")

        brutto_local = totals["brutto"]
        brutto_eur   = r2(brutto_local * rate)
        netto_eur    = r2(brutto_eur / (1 + tax_rate))
        ust_eur      = r2(brutto_eur - netto_eur)

        print(f"\n  Totals (local):")
        for k, v in totals.items():
            if v != 0:
                print(f"    {k}: {v:.4f} {currency}  →  EUR: {r2(v*rate):.2f}")
        print(f"  brutto_eur={brutto_eur:.2f}  netto={netto_eur:.2f}  ust={ust_eur:.2f}")
        print(f"  kto_gut={kto_gut:.2f}  behalten={behalten_val:.2f}")

        erstattung = r2(totals["erstattung"] * rate)
        marktplatz = r2(totals["marktplatz"] * rate)
        commission = r2(totals["commission"] * rate)
        gutschein  = r2(totals["gutschein"]  * rate)
        steuer     = r2(totals["steuer"]     * rate)
        sonstiges  = r2(totals["sonstiges"]  * rate)

        fields = [brutto_eur, erstattung, marktplatz, commission,
                  gutschein, steuer, sonstiges, kto_gut, behalten_val]
        total_fields = r2(sum(fields))
        inner_diff   = r2(total_fields - payout_eur)

        print(f"\n  InnerChk breakdown:")
        print(f"    brutto_eur     = {brutto_eur:>10.2f}")
        print(f"    erstattung     = {erstattung:>10.2f}")
        print(f"    marktplatz     = {marktplatz:>10.2f}")
        print(f"    commission     = {commission:>10.2f}")
        print(f"    gutschein      = {gutschein:>10.2f}")
        print(f"    steuer         = {steuer:>10.2f}")
        print(f"    sonstiges      = {sonstiges:>10.2f}")
        print(f"    kto_gut        = {kto_gut:>10.2f}")
        print(f"    behalten       = {behalten_val:>10.2f}")
        print(f"    ─────────────────────────")
        print(f"    SUM fields     = {total_fields:>10.2f}")
        print(f"    payout_eur     = {payout_eur:>10.2f}")
        print(f"    InnerChk       = {inner_diff:>10.2f}  ← {'OK' if abs(inner_diff)<=0.02 else 'PROBLEM'}")

print("\nDone.")
