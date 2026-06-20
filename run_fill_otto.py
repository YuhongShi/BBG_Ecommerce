import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

src_path = r"D:\Workstation_Yuhong\Yuhong's_File\Python\Buchhaltung\Kontoauszug_Otto.xlsx"
dst_path = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_Mai.xlsx"

# 读取 Kontoauszug_Otto（11列：Nr,Marktplatz,Netto,Ust,Brutto,Ref,Comm,CommR,Pay,PayR,Summe）
wb_src = openpyxl.load_workbook(src_path, data_only=True)
ws_src = wb_src['Otto']
konto_data = {}
for row in ws_src.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    nr = str(int(row[0]))
    konto_data[nr] = row

print("Kontoauszug_Otto 数据:")
for nr, row in konto_data.items():
    print(f"  Nr={nr}, Summe={row[10]}")

wb_dst = openpyxl.load_workbook(dst_path)
ws_otto = wb_dst['OTTO']

# 第一步：清空 OTTO sheet（保留表头），再填入新数据
for row in ws_otto.iter_rows(min_row=2, max_row=ws_otto.max_row):
    for cell in row:
        cell.value = None

appended = []
for nr, src in konto_data.items():
    new_row = [
        int(nr),
        src[1],        # Marktplatz
        src[2],        # Netto_Sale
        src[3],        # Ust
        src[4],        # Brutto_Sale
        src[5],        # Refund(19%)
        src[6],        # Commission(19%)
        src[7],        # Commission_rev(19%)
        src[8],        # Paymentfee(19%)
        src[9],        # Paymentfee_ref(19%)
        None,          # VAT-Korrektur (不在Kontoauszug中)
        round(float(src[10]), 2) if src[10] is not None else None,  # Summe
    ]
    ws_otto.append(new_row)
    appended.append(nr)

print(f"\n第一步：已清空旧数据，写入 Nr={appended} 到 OTTO sheet")

# 第二步：核对 Sheet1 Betrag
ws1 = wb_dst['Sheet1']
otto_summe = {}
for row in ws_otto.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    otto_summe[str(int(row[0]))] = row[11]

print("\n第二步：核对 Sheet1 Betrag vs OTTO Summe")
mismatch = []
for row in ws1.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    zv = str(row[2]) if row[2] else ''
    if 'OTTO' in zv.upper():
        nr = str(int(row[0]))
        betrag = row[3]
        summe = otto_summe.get(nr)
        if summe is not None and abs(float(betrag) - float(summe)) > 0.02:
            mismatch.append((nr, betrag, summe))
            print(f"  ✗ Nr={nr}: Sheet1 Betrag={betrag}, OTTO Summe={summe}")
        else:
            print(f"  ✓ Nr={nr}: Betrag={betrag} == Summe={summe}")

if mismatch:
    print("\n存在不一致，请修正后再继续。")
    sys.exit(1)

# 第三步：填入 Sheet1 Brrutto_Umsatz 和 Andere Gebühren
print("\n第三步：填入 Sheet1 Brrutto_Umsatz 和 Andere Gebühren")
otto_detail = {}
for row in ws_otto.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    nr = str(int(row[0]))
    brutto = row[4] or 0
    refund = row[5] or 0
    comm   = row[6] or 0
    commr  = row[7] or 0
    pay    = row[8] or 0
    payr   = row[9] or 0
    vat    = row[10] or 0
    andere = refund + comm + commr + pay + payr + vat
    otto_detail[nr] = (brutto, andere)

for row in ws1.iter_rows(min_row=2):
    if row[0].value is None:
        continue
    zv = str(row[2].value) if row[2].value else ''
    if 'OTTO' in zv.upper():
        nr = str(int(row[0].value))
        if nr in otto_detail:
            row[4].value = round(otto_detail[nr][0], 2)
            row[5].value = round(otto_detail[nr][1], 2)
            print(f"  Nr={nr}: Brrutto_Umsatz={round(otto_detail[nr][0],2)}, Andere={round(otto_detail[nr][1],2)}")

wb_dst.save(dst_path)
print("\n文件已保存。OTTO 平台处理完成。")
