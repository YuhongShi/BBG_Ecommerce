#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""fill_xpollens.py — Fill XPOLLENS (Leroy Merlin) rows in Sheet1.

For each billing-cycle CSV of the target month, computes:
  Brutto_Umsatz  = sum(Order amount + tax + shipping) + sum(refunds)
  Andere Gebühren = sum(Commission + Subscription fee)
  Betrag         = -Payment row amount

Matches to Sheet1 XPOLLENS entries by Betrag (exact), then writes
Sheet1 cols: 5=Brrutto_Umsatz, 6=Andere Gebühren, 11=Steuersatz(0), 12=Marktplatz.
"""

import sys, os, re, argparse, csv
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

parser = argparse.ArgumentParser()
parser.add_argument('--umsatz', required=True, help='Path to Umsatzsteuer .xlsx')
parser.add_argument('--year',   required=True, type=int, help='Year  (e.g. 2026)')
parser.add_argument('--month',  required=True, type=int, help='Month (e.g. 4)')
args = parser.parse_args()

XPOLLENS_BASE = r'D:\Workstation_Yuhong\BBG\电商\网买做账\XPOLLENS'


def parse_amt(s):
    s = s.strip()
    return float(s.replace(',', '.')) if s else 0.0


def process_csv(path):
    """Return (betrag, brutto, andere) or None on parse failure."""
    with open(path, encoding='latin-1') as fh:
        rows = list(csv.reader(fh, delimiter=';'))
    if not rows:
        return None
    hdr = rows[0]
    try:
        type_col   = hdr.index('Type')
        amount_col = hdr.index('Amount')
    except ValueError:
        return None

    sale = refund = commission = subscription = payment = 0.0
    for r in rows[1:]:
        if len(r) <= max(type_col, amount_col):
            continue
        t   = r[type_col].strip()
        amt = parse_amt(r[amount_col])

        if t in ('Order amount', 'Order amount tax',
                 'Shipping charges', 'Shipping tax'):
            sale += amt
        elif t in ('Order amount refund', 'Order amount tax refund',
                   'Shipping charge refund', 'Shipping tax refund'):
            refund += amt
        elif t in ('Commission', 'Commission (excl, tax)', 'Commission refund'):
            commission += amt
        elif t == 'Subscription fee':
            subscription += amt
        elif t == 'Payment':
            payment += amt

    betrag = round(-payment, 2)
    brutto = round(sale + refund, 2)
    andere = round(commission + subscription, 2)
    return betrag, brutto, andere


# ── Find CSVs for the target month ────────────────────────────────────────────
yy      = args.year % 100
mm      = args.month
pattern = re.compile(rf'{yy}\.{mm}\.\d+回款明细\.csv$')

csv_records = []  # (betrag, brutto, andere, filename)
for dirpath, _, filenames in os.walk(XPOLLENS_BASE):
    for fname in filenames:
        if pattern.search(fname):
            result = process_csv(os.path.join(dirpath, fname))
            if result:
                csv_records.append((*result, fname))

csv_records.sort(key=lambda x: x[0], reverse=True)
print(f"Found {len(csv_records)} XPOLLENS CSV(s) for {args.year}.{mm:02d}:")
for betrag, brutto, andere, fname in csv_records:
    print(f"  {fname}  →  Betrag={betrag:.2f}  Brutto={brutto:.2f}  Andere={andere:.2f}")

# ── Load Umsatzsteuer ──────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(args.umsatz)
ws = wb['Sheet1']

s1_rows = []
for r in range(2, ws.max_row + 1):
    z = ws.cell(r, 3).value
    b = ws.cell(r, 4).value
    if z and 'XPOLLENS' in str(z) and b:
        s1_rows.append({'row': r, 'betrag': float(b), 'matched': False})
print(f"\nSheet1 XPOLLENS rows: {len(s1_rows)}")

# ── Match by Betrag ────────────────────────────────────────────────────────────
matches = []
for betrag, brutto, andere, fname in csv_records:
    best = next(
        (e for e in s1_rows if not e['matched'] and abs(e['betrag'] - betrag) < 0.02),
        None
    )
    if best:
        best['matched'] = True
        matches.append((best, betrag, brutto, andere, fname))
        print(f"  MATCH  {fname:<50s}  Betrag={betrag:.2f}  →  Row {best['row']}")
    else:
        print(f"  SKIP   {fname:<50s}  Betrag={betrag:.2f}  (no Sheet1 match)")

unmatched = [e for e in s1_rows if not e['matched']]
if unmatched:
    print(f"\n⚠ Unmatched Sheet1 XPOLLENS entries ({len(unmatched)}):")
    for e in unmatched:
        print(f"  Row={e['row']}  Betrag={e['betrag']:.2f}")
else:
    print("\n✓ All Sheet1 XPOLLENS entries matched.")

# ── Write Sheet1 ───────────────────────────────────────────────────────────────
if not ws.cell(1, 11).value:
    ws.cell(1, 11).value = 'Steuersatz'
if not ws.cell(1, 12).value:
    ws.cell(1, 12).value = 'Marktplatz'

print("\n=== Sheet1 updates ===")
print(f"{'Row':>4}  {'Brutto':>10}  {'Andere':>10}  {'Betrag':>10}  Check")
print("─" * 50)
for s1, betrag, brutto, andere, fname in matches:
    r = s1['row']
    ws.cell(r,  5).value = brutto or None
    ws.cell(r,  6).value = andere or None
    ws.cell(r, 11).value = 0.0
    ws.cell(r, 12).value = 'Leroymerlin'
    ok = '✓' if abs(brutto + andere - betrag) < 0.02 else '⚠'
    print(f"{r:>4}  {brutto:>10.2f}  {andere:>10.2f}  {betrag:>10.2f}  {ok}")

wb.save(args.umsatz)
print(f"\n✓ Saved: {args.umsatz}")
print(f"  {len(matches)} XPOLLENS rows updated in Sheet1.")
