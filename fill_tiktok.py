#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""fill_tiktok.py — Fill TikTok rows in Sheet1.

For each Paid payment in the TikTok settlement xlsx, computes:
  Brrutto_Umsatz  = Net sales (sum across all statements for that Payment ID)
  Andere Gebühren = Shipping + Fees + Adjustments
  Betrag          = Payment amount

Matches to Sheet1 TikTok entries by Betrag (exact), then writes
Sheet1 cols: 5=Brrutto_Umsatz, 6=Andere Gebühren, 11=Steuersatz(0), 12=Marktplatz.
"""

import sys, os, re, argparse, glob
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

parser = argparse.ArgumentParser()
parser.add_argument('--umsatz',  required=True, help='Path to Umsatzsteuer .xlsx')
parser.add_argument('--tiktok',  required=False, help='Path to TikTok settlement .xlsx (optional, auto-detected if omitted)')
parser.add_argument('--year',    required=False, type=int, help='Year for auto-detection (e.g. 2026)')
parser.add_argument('--month',   required=False, type=int, help='Month for auto-detection (e.g. 5)')
args = parser.parse_args()

TIKTOK_BASE = r'D:\Workstation_Yuhong\BBG\电商\网买做账\Tiktok'


def parse_amt(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', '.'))
    except ValueError:
        return 0.0


# ── Locate TikTok file ────────────────────────────────────────────────────────
tiktok_path = args.tiktok
if not tiktok_path:
    if not args.year or not args.month:
        print("ERROR: provide --tiktok path, or both --year and --month for auto-detection.")
        sys.exit(1)
    # search in year directory and all monthly subdirs
    year_dir = os.path.join(TIKTOK_BASE, str(args.year))
    candidates = glob.glob(os.path.join(year_dir, '**', '*.xlsx'), recursive=True)
    candidates = [f for f in candidates if 'tiktok' in os.path.basename(f).lower()
                                        and '订单' in os.path.basename(f)]
    if not candidates:
        print(f"ERROR: No TikTok settlement xlsx found under {year_dir}")
        sys.exit(1)
    # prefer file whose name contains the target month
    yy = args.year % 100
    mm = args.month
    preferred = [f for f in candidates if f'{yy}.{mm}.' in os.path.basename(f)]
    tiktok_path = preferred[0] if preferred else candidates[0]

print(f"TikTok file: {tiktok_path}")

# ── Read TikTok xlsx ──────────────────────────────────────────────────────────
wb_tt = openpyxl.load_workbook(tiktok_path, data_only=True)

# Payments sheet: get all Paid payments
ws_pay = wb_tt['Payments']
pay_headers = [c.value for c in ws_pay[1]]
pay_id_col     = pay_headers.index('Payment ID')
pay_amt_col    = pay_headers.index('Payment amount')
pay_status_col = pay_headers.index('Status')

paid_payments = {}  # payment_id → amount
for row in ws_pay.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    pid    = str(row[pay_id_col]).strip()
    status = str(row[pay_status_col]).strip()
    amt    = parse_amt(row[pay_amt_col])
    if status == 'Paid' and pid:
        paid_payments[pid] = round(amt, 2)

print(f"\nPaid payments found: {len(paid_payments)}")
for pid, amt in paid_payments.items():
    print(f"  Payment ID {pid}: {amt:.2f} EUR")

# Statements sheet: aggregate by Payment ID
ws_stmt = wb_tt['Statements']
stmt_headers = [c.value for c in ws_stmt[1]]
stmt_pid_col   = stmt_headers.index('Payment ID')
stmt_net_col   = stmt_headers.index('Net sales')
stmt_ship_col  = stmt_headers.index('Shipping')
stmt_fees_col  = stmt_headers.index('Fees')
stmt_adj_col   = stmt_headers.index('Adjustments')

aggregated = {}  # payment_id → {net, andere}
for row in ws_stmt.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    pid = str(row[stmt_pid_col]).strip()
    if pid not in paid_payments:
        continue
    net   = parse_amt(row[stmt_net_col])
    ship  = parse_amt(row[stmt_ship_col])
    fees  = parse_amt(row[stmt_fees_col])
    adj   = parse_amt(row[stmt_adj_col])
    if pid not in aggregated:
        aggregated[pid] = {'net': 0.0, 'andere': 0.0}
    aggregated[pid]['net']    += net
    aggregated[pid]['andere'] += ship + fees + adj

payments_data = []  # (betrag, brutto, andere, payment_id)
for pid, amt in paid_payments.items():
    if pid not in aggregated:
        print(f"  [警告] Payment {pid} has no matching Statements rows, skipping.")
        continue
    brutto = round(aggregated[pid]['net'],    2)
    andere = round(aggregated[pid]['andere'], 2)
    payments_data.append((amt, brutto, andere, pid))

print(f"\nSettlement summary:")
print(f"{'Payment ID':<25} {'Betrag':>8} {'Brutto':>10} {'Andere':>10} {'Check'}")
print('─' * 65)
for betrag, brutto, andere, pid in payments_data:
    chk = '✓' if abs(brutto + andere - betrag) < 0.02 else '⚠'
    print(f"{pid:<25} {betrag:>8.2f} {brutto:>10.2f} {andere:>10.2f}  {chk}")

# ── Load Umsatzsteuer ─────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(args.umsatz)
ws = wb['Sheet1']

s1_rows = []
for r in range(2, ws.max_row + 1):
    z = ws.cell(r, 3).value
    b = ws.cell(r, 4).value
    if z and str(z).strip().lower() == 'tiktok' and b:
        s1_rows.append({'row': r, 'betrag': float(b), 'matched': False})
print(f"\nSheet1 TikTok rows: {len(s1_rows)}")

# ── Match by Betrag ───────────────────────────────────────────────────────────
matches = []
for betrag, brutto, andere, pid in payments_data:
    best = next(
        (e for e in s1_rows if not e['matched'] and abs(e['betrag'] - betrag) < 0.02),
        None
    )
    if best:
        best['matched'] = True
        matches.append((best, betrag, brutto, andere, pid))
        print(f"  MATCH  Payment {pid}  Betrag={betrag:.2f}  →  Row {best['row']}")
    else:
        print(f"  SKIP   Payment {pid}  Betrag={betrag:.2f}  (no Sheet1 match)")

unmatched = [e for e in s1_rows if not e['matched']]
if unmatched:
    print(f"\n⚠ Unmatched Sheet1 TikTok entries ({len(unmatched)}):")
    for e in unmatched:
        print(f"  Row={e['row']}  Betrag={e['betrag']:.2f}")
else:
    print("\n✓ All Sheet1 TikTok entries matched.")

# ── Write Sheet1 ──────────────────────────────────────────────────────────────
if not ws.cell(1, 11).value:
    ws.cell(1, 11).value = 'Steuersatz'
if not ws.cell(1, 12).value:
    ws.cell(1, 12).value = 'Marktplatz'

print("\n=== Sheet1 updates ===")
print(f"{'Row':>4}  {'Brutto':>10}  {'Andere':>10}  {'Betrag':>10}  Check")
print("─" * 50)
for s1, betrag, brutto, andere, pid in matches:
    r = s1['row']
    ws.cell(r,  5).value = brutto or None
    ws.cell(r,  6).value = andere or None
    ws.cell(r, 11).value = 0.0
    ws.cell(r, 12).value = 'TikTok'
    ok = '✓' if abs(brutto + andere - betrag) < 0.02 else '⚠'
    print(f"{r:>4}  {brutto:>10.2f}  {andere:>10.2f}  {betrag:>10.2f}  {ok}")

wb.save(args.umsatz)
print(f"\n✓ Saved: {args.umsatz}")
print(f"  {len(matches)} TikTok rows updated in Sheet1.")
