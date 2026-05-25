# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

KONTOAUSZUG = r"D:\Workstation_Yuhong\Yuhong's_File\Python\Buchhaltung\Kontoauszug_Ebay.xlsx"
UMSATZ = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_März.xlsx"

# All Kontoauszug Ebay_Umsatz rows
wb_k = openpyxl.load_workbook(KONTOAUSZUG, data_only=True)
ws_k = wb_k["Ebay_Umsatz"]
print(f"=== Kontoauszug Ebay_Umsatz (all {ws_k.max_row-1} data rows) ===")
for i, row in enumerate(ws_k.iter_rows(min_row=1, values_only=True)):
    if any(v is not None for v in row):
        print(f"  row{i+1}: {row}")

# Umsatzsteuer Ebay sheet - all rows
wb_u = openpyxl.load_workbook(UMSATZ, data_only=True)
ws_ebay = wb_u["Ebay"]
print(f"\n=== Umsatzsteuer Ebay sheet (all rows, max_row={ws_ebay.max_row}) ===")
for i, row in enumerate(ws_ebay.iter_rows(min_row=1, values_only=True)):
    if any(v is not None for v in row):
        print(f"  row{i+1}: {row}")
    elif i == 0:
        print(f"  row{i+1}: (header empty?)")
