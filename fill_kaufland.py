#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""fill_kaufland.py — Fill Kaufland sheet + Sheet1 in Umsatzsteuer file.

Column structure (14 cols):
  1=Nr, 2=Marktplatz, 3=Netto_Sale, 4=Ust, 5=Brutto_Sale,
  6=Erstattung(19%), 7=Marktpaltzgebühr(19%), 8=Commission(19%),
  9=Gutscheingebühr(19%), 10=Steuer_Ohne_Gebühr(19%), 11=Sonstiges(19%),
  12=KTO_Guthaben, 13=Behalten_in_KTO, 14=Summe

Sheet1 target:
  Col 5=Brrutto_Umsatz, Col 6=Andere Gebühren,
  Col 7=KTO_Guthaben, Col 8=Behalten_in_KTO
"""

import sys, re, argparse, os
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

parser = argparse.ArgumentParser()
parser.add_argument("--umsatz",        required=True)
parser.add_argument("--kaufland",      required=True,
                    help="Current-period Kaufland file (施总)")
parser.add_argument("--prev-kaufland", required=False, default=None,
                    help="Previous-period Kaufland file (needed for DE cross-month tail)")
args = parser.parse_args()

KAUFLAND_FILE = args.kaufland
PREV_KL_FILE  = args.prev_kaufland
UMSATZ_FILE   = args.umsatz

COUNTRY_CFG = {
    "德国站":     {"currency": "EUR", "behalten": True},
    "奥地利站":   {"currency": "EUR", "behalten": False},
    "法国站":     {"currency": "EUR", "behalten": False},
    "斯洛伐克站": {"currency": "EUR", "behalten": False},
    "意大利站":   {"currency": "EUR", "behalten": False},
    "捷克站":     {"currency": "CZK", "behalten": False},
    "波兰站":     {"currency": "PLN", "behalten": False},
}
COUNTRY_S1_LABELS = {
    "德国站":     ["Kaufland", "Kaufland-DE"],
    "奥地利站":   ["Kaufland", "Kaufland-AT"],
    "法国站":     ["Kaufland", "Kaufland-FR"],
    "斯洛伐克站": ["Kaufland", "Kaufland-SK"],
    "意大利站":   ["Kaufland", "Kaufland-IT"],
    "捷克站":     ["KTO", "Kaufland-CZ"],
    "波兰站":     ["IBAN", "IBAN(Unbekannt)", "Kaufland-PL"],
}
COUNTRY_MARKT = {
    "德国站":     "Kaufland-DE",
    "奥地利站":   "Kaufland-AT",
    "法国站":     "Kaufland-FR",
    "斯洛伐克站": "Kaufland-SK",
    "意大利站":   "Kaufland-IT",
    "捷克站":     "Kaufland-CZ",
    "波兰站":     "Kaufland-PL",
}
COUNTRY_TAX_RATE = {
    "德国站":     0.19,
    "奥地利站":   0.20,
    "法国站":     0.20,
    "斯洛伐克站": 0.23,  # raised from 20% to 23% in Jan 2025
    "意大利站":   0.22,
    "捷克站":     0.21,
    "波兰站":     0.23,
}
ALL_KL_LABELS = {
    "Kaufland", "Kaufland-DE", "Kaufland-AT", "Kaufland-FR",
    "Kaufland-SK", "Kaufland-IT", "Kaufland-CZ", "Kaufland-PL",
    "KTO", "IBAN", "IBAN(Unbekannt)",
}
FUZZY_TOL = 0.06

# ── helpers ────────────────────────────────────────────────────────────────
def parse_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try: return float(s)
    except: return 0.0

def r2(v): return round(v, 2)

def extract_ym(sheet_name):
    """Return (year, month) int tuple from '26年3月...' style name, or None."""
    m = re.search(r'(\d+)年(\d+)月', sheet_name)
    return (int(m.group(1)), int(m.group(2))) if m else None

def read_sheet_rows(ws):
    """Locate header row (booking_date in col 1) and return list of row dicts."""
    hdr = next(
        (r for r in range(1, 12) if str(ws.cell(r, 1).value or "").strip() == "booking_date"),
        None
    )
    if hdr is None:
        return []
    rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        bt = ws.cell(r, 2).value
        if bt is None:
            continue
        rows.append({
            "bt":      str(bt).strip(),
            "amount":  parse_num(ws.cell(r, 5).value),
            "balance": parse_num(ws.cell(r, 6).value),
            "sgross":  parse_num(ws.cell(r, 9).value),
            "fgross":  parse_num(ws.cell(r, 13).value),
            "vat_pct": parse_num(ws.cell(r, 12).value),  # fee_vat_% col
        })
    return rows

def get_prev_tail(ws):
    """Return (last_payout_balance_local, tail_rows) from a previous-period sheet.

    last_payout_balance_local is the balance AFTER the final payout in that sheet
    (= KTO_Guthaben for the first current-period payout).
    tail_rows are all transaction rows that follow the last payout.
    """
    rows = read_sheet_rows(ws)
    if not rows:
        return 0.0, []
    last_payout_idx = None
    for i, row in enumerate(rows):
        if row["bt"] == "Payout":
            last_payout_idx = i
    if last_payout_idx is None:
        # No payout in prev sheet — entire sheet is a tail; opening balance = first row's balance minus amount
        opening = rows[0]["balance"] - rows[0]["amount"]
        return opening, rows
    return rows[last_payout_idx]["balance"], rows[last_payout_idx + 1:]

def classify_tx(bt, amt, sg, fg):
    """Classify one transaction row. Returns list of (field_name, value) tuples.

    "Netto X" and "Umsatzsteuer X" are the same fee split into net + tax;
    both go to the same column so they sum to the gross (19%) amount.
    Zero-tax fees (e.g. "Fees for cancelled orders") go to Steuer_Ohne_Gebühr.
    """
    if bt.startswith("Freigabe") or bt.startswith("Release order"):
        return [("brutto", sg), ("commission", -fg)]
    if (bt.startswith("Erstattung") or bt.startswith("Partial Refund")
            or bt.startswith("Storno")):
        return [("erstattung", amt)]

    # Zero-tax fees: no corresponding Umsatzsteuer line → Steuer_Ohne_Gebühr
    bt_lo = bt.lower()
    if bt_lo.startswith("fees for cancelled") or bt_lo.startswith("gebühren für stornierte"):
        return [("steuer", amt)]

    # Strip "Netto " / "Net " / "Umsatzsteuer " prefix to get canonical description,
    # then classify Netto and Umsatzsteuer entries into the same column.
    canonical = bt
    for prefix in ("Umsatzsteuer ", "Netto ", "Net "):
        if canonical.startswith(prefix):
            canonical = canonical[len(prefix):]
            break
    canonical_lo = canonical.lower()

    if "grundgebühr" in canonical_lo:
        return [("marktplatz", amt)]
    if "gutschein" in canonical_lo or "voucher" in canonical_lo:
        return [("gutschein", amt)]
    # Sponsored ads, EPR fees, and other unrecognised entries → Sonstiges
    return [("sonstiges", amt)]

# ── Step 1: load files ─────────────────────────────────────────────────────
print("Loading Kaufland raw data…")
wb_kl = openpyxl.load_workbook(KAUFLAND_FILE, data_only=True)

if PREV_KL_FILE:
    print("Loading previous-period Kaufland file…")
    wb_prev = openpyxl.load_workbook(PREV_KL_FILE, data_only=True)
else:
    wb_prev = None

print("Loading Umsatzsteuer file…")
wb_us = openpyxl.load_workbook(UMSATZ_FILE, data_only=True)
ws_s1   = wb_us["Sheet1"]
ws_kauf = wb_us["Kaufland"]

# ── Step 2: exchange rates ─────────────────────────────────────────────────
rates = {"EUR": 1.0, "PLN": 0.2356, "CZK": 0.0411}
ws_sum = wb_kl["合计应结算"]
for row in range(14, 22):
    label = str(ws_sum.cell(row, 1).value or "")
    rtext = str(ws_sum.cell(row, 3).value or "")
    m = re.search(r"\*\s*(\d+\.?\d+)", rtext)
    if m:
        rate = float(m.group(1))
        if "PLN" in label: rates["PLN"] = rate
        if "CZK" in label: rates["CZK"] = rate
print(f"Exchange rates: PLN={rates['PLN']}, CZK={rates['CZK']}")

# ── Step 3: map sheets in current file by country and year-month ───────────
# country → sorted list of (ym_tuple, sheet_name), latest last
country_sheet_map = {}
for sn in wb_kl.sheetnames:
    country = next((k for k in COUNTRY_CFG if k in sn), None)
    if not country:
        continue
    ym = extract_ym(sn)
    if ym is None:
        continue
    country_sheet_map.setdefault(country, []).append((ym, sn))
for country in country_sheet_map:
    country_sheet_map[country].sort(key=lambda x: x[0])

# ── Step 4: build Sheet1 lookup ────────────────────────────────────────────
s1_entries = []
for r in range(2, ws_s1.max_row + 1):
    nr     = ws_s1.cell(r, 1).value
    z      = ws_s1.cell(r, 3).value
    betrag = ws_s1.cell(r, 4).value
    if not nr or not z or not betrag:
        continue
    zs = str(z)
    if zs in ALL_KL_LABELS:
        s1_entries.append({"nr": nr, "row": r, "label": zs, "betrag": float(betrag), "matched": False})
print(f"Sheet1 Kaufland/KTO/IBAN rows: {len(s1_entries)}")

# ── Step 5: process each country → collect payout records ─────────────────
all_payouts = []

for country, cfg in COUNTRY_CFG.items():
    currency = cfg["currency"]
    rate     = rates[currency]

    sheets_for_country = country_sheet_map.get(country, [])
    if not sheets_for_country:
        print(f"  [SKIP] No sheet found for {country} in current file")
        continue

    # Current-period sheet = latest
    current_sn = sheets_for_country[-1][1]
    ws_cur = wb_kl[current_sn]

    # Locate previous-period tail
    prev_bal_local = 0.0
    prev_tail_rows = []

    if len(sheets_for_country) >= 2:
        # Prev-period sheet is embedded in the same file (non-DE multi-month)
        prev_sn = sheets_for_country[-2][1]
        prev_bal_local, prev_tail_rows = get_prev_tail(wb_kl[prev_sn])
        print(f"  [{country}] prev sheet '{prev_sn}': "
              f"last_payout_bal={prev_bal_local:.2f}{currency}, tail_rows={len(prev_tail_rows)}")
    elif country == "德国站":
        if wb_prev is None:
            print(f"  [WARN] 德国站 needs --prev-kaufland for cross-month tail; skipping prev tail")
        else:
            # Find the DE sheet in the prev file (typically named something like '26年2月' or '26年2月(德国站)')
            de_prev_sn = next(
                (sn for sn in wb_prev.sheetnames if "德国站" in sn or
                 (extract_ym(sn) is not None and not any(k in sn for k in COUNTRY_CFG if k != "德国站"))),
                None
            )
            if de_prev_sn is None:
                # Fallback: pick first non-summary sheet with year-month
                de_prev_sn = next((sn for sn in wb_prev.sheetnames if extract_ym(sn) is not None
                                   and "合计" not in sn), None)
            if de_prev_sn:
                prev_bal_local, prev_tail_rows = get_prev_tail(wb_prev[de_prev_sn])
                print(f"  [{country}] prev sheet '{de_prev_sn}' (from prev file): "
                      f"last_payout_bal={prev_bal_local:.2f}{currency}, tail_rows={len(prev_tail_rows)}")
            else:
                print(f"  [WARN] Could not locate DE sheet in prev Kaufland file")
    else:
        # Non-DE country without embedded prev sheet.
        # Read the tail rows from wb_prev (transactions after the last prev-period payout).
        # These carry the account's opening balance into the first current-period group.
        if wb_prev is not None:
            # Pick the sheet with the LATEST year-month for this country in wb_prev.
            # (prev file may contain both current and embedded-prev sheets; we want the latest.)
            non_de_candidates = [
                (extract_ym(sn), sn) for sn in wb_prev.sheetnames
                if country in sn and extract_ym(sn) is not None
            ]
            non_de_prev_sn = max(non_de_candidates, key=lambda x: x[0])[1] if non_de_candidates else None
            if non_de_prev_sn:
                prev_bal_local, prev_tail_rows = get_prev_tail(wb_prev[non_de_prev_sn])
                print(f"  [{country}] prev tail '{non_de_prev_sn}' (prev file): "
                      f"last_payout_bal={prev_bal_local:.2f}{currency}, tail_rows={len(prev_tail_rows)}")

    # Read current-period rows
    cur_rows = read_sheet_rows(ws_cur)
    if not cur_rows:
        print(f"  [SKIP] No data rows in {current_sn}")
        continue

    # Split current rows into groups ending at each Payout
    groups = []
    head   = list(prev_tail_rows)  # start with prev tail
    for row in cur_rows:
        if row["bt"] == "Payout":
            groups.append({"txns": head, "payout": row})
            head = []
        else:
            head.append(row)
    # 'head' now contains tail transactions of current period (go to next month payout)

    prev_bal = prev_bal_local  # track balance in local currency

    # Derive tax rate from first Freigabe row with non-zero fee_vat_%
    tax_rate = 0.19  # fallback
    for row in prev_tail_rows + cur_rows:
        bt = row["bt"]
        if (bt.startswith("Freigabe") or bt.startswith("Release order")) and row["vat_pct"] > 0:
            tax_rate = row["vat_pct"] / 100
            break

    for g in groups:
        txns         = g["txns"]
        payout       = g["payout"]
        payout_local = abs(payout["amount"])
        payout_eur   = r2(payout_local * rate)

        old_prev_bal = prev_bal
        kto_gut      = r2(prev_bal * rate)
        behalten_val = r2(-payout["balance"] * rate)  # negative: amount kept back (deduction sign)
        prev_bal     = payout["balance"]              # 0 for non-DE, retained for DE

        if not txns:
            # Pure-balance withdrawal: no underlying sale transactions
            rec = dict(
                sheet=current_sn, country=country, currency=currency, tax_rate=tax_rate,
                brutto=0.0, netto=0.0, ust=0.0,
                erstattung=0.0, marktplatz=0.0, commission=0.0,
                gutschein=0.0, steuer=0.0, sonstiges=0.0,
                kto_gut=kto_gut, behalten=behalten_val,
                kto_gut_local=old_prev_bal, behalten_local=-payout["balance"],
                rate_used=rate,
                summe_local=r2(payout_local), summe_eur=payout_eur,
                pure_balance=True,
            )
        else:
            totals = {"brutto": 0.0, "erstattung": 0.0, "marktplatz": 0.0,
                      "commission": 0.0, "gutschein": 0.0, "steuer": 0.0,
                      "sonstiges": 0.0}
            for tx in txns:
                for field, val in classify_tx(tx["bt"], tx["amount"], tx["sgross"], tx["fgross"]):
                    totals[field] += val

            brutto_local = totals["brutto"]
            brutto_eur   = r2(brutto_local * rate)
            netto_eur    = r2(brutto_eur / (1 + tax_rate))
            ust_eur      = r2(brutto_eur - netto_eur)

            rec = dict(
                sheet=current_sn, country=country, currency=currency, tax_rate=tax_rate,
                brutto=brutto_eur,
                netto=netto_eur,
                ust=ust_eur,
                erstattung=r2(totals["erstattung"] * rate),
                marktplatz=r2(totals["marktplatz"] * rate),
                commission=r2(totals["commission"] * rate),
                gutschein=r2(totals["gutschein"]   * rate),
                steuer=r2(totals["steuer"]         * rate),
                sonstiges=r2(totals["sonstiges"]   * rate),
                kto_gut=kto_gut,
                behalten=behalten_val,
                totals_local=dict(totals),
                kto_gut_local=old_prev_bal,
                behalten_local=-payout["balance"],
                rate_used=rate,
                summe_local=r2(payout_local),
                summe_eur=payout_eur,
                pure_balance=False,
            )
        all_payouts.append(rec)

# ── Step 5b: Prev-file last payouts (cross-month arrivals) ────────────────────
# When the current file has no embedded prev-month sheet for a given country,
# the previous month's LAST payout may have arrived in the current month's bank
# account.  Process it from the prev file so it can be matched to Sheet1.
if wb_prev:
    # Load prev-month pending file to know which payouts were actually SKIP'd.
    # Only those should be re-attempted in Step 5b; ones that were already
    # matched in the prev month must not appear as "pending" in the current month.
    umsatz_dir    = os.path.dirname(os.path.abspath(UMSATZ_FILE))
    monat_m_curr  = re.search(r'Umsatzsteuer_(.+)\.xlsx', os.path.basename(UMSATZ_FILE))
    current_monat = monat_m_curr.group(1) if monat_m_curr else None
    pp_candidates = []
    for _fn in os.listdir(umsatz_dir):
        _pm = re.match(r'Kaufland_pending_(.+)\.txt', _fn)
        if _pm and _pm.group(1) != current_monat:
            pp_candidates.append(_fn)
    prev_pending_set = None  # None = no filtering (prev file not found → first run)
    if pp_candidates:
        # Most recently modified = immediately previous month
        pp_candidates.sort(
            key=lambda f: os.path.getmtime(os.path.join(umsatz_dir, f)), reverse=True)
        pp_path = os.path.join(umsatz_dir, pp_candidates[0])
        prev_pending_set = set()
        with open(pp_path, encoding="utf-8") as _pp:
            for _line in _pp:
                _line = _line.rstrip()
                if not _line or _line.startswith('#') or '─' in _line or _line.startswith('国家'):
                    continue
                _parts = _line.split()
                if not _parts:
                    continue
                _ctry = _parts[0]
                # Non-EUR line: "AMOUNT CZK/PLN  (X.XX EUR)" → key by local amount
                _m_loc = re.search(r'(\d+\.\d+)\s+(CZK|PLN)\s+\(', _line)
                # EUR line: "AMOUNT EUR" at end → key by EUR amount
                _m_eur = re.search(r'(\d+\.\d+)\s+EUR\s*$', _line)
                if _m_loc:
                    prev_pending_set.add((_ctry, round(float(_m_loc.group(1)), 2)))
                elif _m_eur:
                    prev_pending_set.add((_ctry, round(float(_m_eur.group(1)), 2)))
        print(f"  Prev pending: {pp_candidates[0]}  ({len(prev_pending_set)} entries to re-check)")
    else:
        print(f"  Prev pending: (none found — no filtering for Step 5b)")

    # Build country→sheets map for prev file
    prev_csm = {}
    for sn in wb_prev.sheetnames:
        c = next((k for k in COUNTRY_CFG if k in sn), None)
        if not c:
            continue
        ym = extract_ym(sn)
        if ym:
            prev_csm.setdefault(c, []).append((ym, sn))
    for c in prev_csm:
        prev_csm[c].sort(key=lambda x: x[0])

    # Read prev-file exchange rates
    prev_rates = dict(rates)  # start with current rates as fallback
    if "合计应结算" in wb_prev.sheetnames:
        ws_ps = wb_prev["合计应结算"]
        for row in range(14, 22):
            lbl   = str(ws_ps.cell(row, 1).value or "")
            rtxt  = str(ws_ps.cell(row, 3).value or "")
            m_r   = re.search(r"\*\s*(\d+\.?\d+)", rtxt)
            if m_r:
                rv = float(m_r.group(1))
                if "PLN" in lbl: prev_rates["PLN"] = rv
                if "CZK" in lbl: prev_rates["CZK"] = rv

    for country, cfg in COUNTRY_CFG.items():
        # Skip German — already handled via get_prev_tail in Step 5 for tail rows;
        # German still needs its own last-payout group added here.
        sheets = prev_csm.get(country, [])
        if not sheets:
            continue

        # Only process countries whose current-file sheets have no embedded prev month
        # (i.e., len == 1 and not matched yet), PLUS German station's last payout.
        cur_sheets = country_sheet_map.get(country, [])
        has_embedded_prev = len(cur_sheets) >= 2
        if has_embedded_prev:
            continue  # prev tail already handled in Step 5

        currency   = cfg["currency"]
        rate       = prev_rates[currency]
        latest_sn  = sheets[-1][1]
        ws_pc      = wb_prev[latest_sn]
        prev_rows  = read_sheet_rows(ws_pc)
        if not prev_rows:
            continue

        payout_indices = [i for i, r in enumerate(prev_rows) if r["bt"] == "Payout"]
        if not payout_indices:
            continue

        # Detect tax rate once per country (not per payout)
        tax_rate = COUNTRY_TAX_RATE[country]
        for row in prev_rows:
            if (row["bt"].startswith("Freigabe") or row["bt"].startswith("Release order")) and row["vat_pct"] > 0:
                tax_rate = row["vat_pct"] / 100
                break

        # Iterate over all payouts in the prev file.
        # With prev_pending_set: include only entries that were SKIP'd in prev month.
        # Without prev_pending_set (first run): include only the last payout.
        for k, cur_idx in enumerate(payout_indices):
            prev_idx     = payout_indices[k - 1] if k > 0 else None
            payout_row   = prev_rows[cur_idx]
            payout_local = abs(payout_row["amount"])
            payout_eur   = r2(payout_local * rate)

            # Filter: local amount for CZK/PLN (EUR equiv varies by rate), EUR otherwise
            _pp_key = (country, round(payout_local, 2)) if currency != "EUR" else (country, payout_eur)
            if prev_pending_set is not None:
                if _pp_key not in prev_pending_set:
                    continue  # already matched in prev month — not pending
            else:
                if cur_idx != payout_indices[-1]:
                    continue  # first-run fallback: only last payout

            start_idx     = (prev_idx + 1) if prev_idx is not None else 0
            group_txns    = prev_rows[start_idx:cur_idx]
            kto_gut_local = prev_rows[prev_idx]["balance"] if prev_idx is not None else 0.0
            kto_gut       = r2(kto_gut_local * rate)
            behalten_val  = r2(payout_row["balance"] * rate)

            if group_txns:
                totals = {"brutto": 0.0, "erstattung": 0.0, "marktplatz": 0.0,
                          "commission": 0.0, "gutschein": 0.0, "steuer": 0.0,
                          "sonstiges": 0.0}
                for tx in group_txns:
                    for field, val in classify_tx(tx["bt"], tx["amount"], tx["sgross"], tx["fgross"]):
                        totals[field] += val
                brutto_local = totals["brutto"]
                brutto_eur   = r2(brutto_local * rate)
                netto_eur    = r2(brutto_eur / (1 + tax_rate))
                ust_eur      = r2(brutto_eur - netto_eur)
                rec = dict(
                    sheet=f"[prev]{latest_sn}", country=country, currency=currency, tax_rate=tax_rate,
                    brutto=brutto_eur, netto=netto_eur, ust=ust_eur,
                    erstattung=r2(totals["erstattung"] * rate),
                    marktplatz=r2(totals["marktplatz"] * rate),
                    commission=r2(totals["commission"] * rate),
                    gutschein=r2(totals["gutschein"]   * rate),
                    steuer=r2(totals["steuer"]         * rate),
                    sonstiges=r2(totals["sonstiges"]   * rate),
                    kto_gut=kto_gut, behalten=behalten_val,
                    totals_local=dict(totals),
                    kto_gut_local=kto_gut_local,
                    behalten_local=-payout_row["balance"],
                    rate_used=rate,
                    summe_local=r2(payout_local), summe_eur=payout_eur,
                    pure_balance=False,
                )
            else:
                rec = dict(
                    sheet=f"[prev]{latest_sn}", country=country, currency=currency, tax_rate=tax_rate,
                    brutto=0.0, netto=0.0, ust=0.0,
                    erstattung=0.0, marktplatz=0.0, commission=0.0,
                    gutschein=0.0, steuer=0.0, sonstiges=0.0,
                    kto_gut=kto_gut, behalten=behalten_val,
                    kto_gut_local=kto_gut_local,
                    behalten_local=-payout_row["balance"],
                    rate_used=rate,
                    summe_local=r2(payout_local), summe_eur=payout_eur,
                    pure_balance=True,
                )
            all_payouts.append(rec)
            print(f"  [PREV] {country:10s} {latest_sn:30s}  payout={payout_eur:.2f}EUR")

print(f"Total payout records: {len(all_payouts)}")

# ── Step 6: match payouts → Sheet1 entries ─────────────────────────────────
# Sort by summe_eur descending so larger (more uniquely identifiable) payouts
# claim their Sheet1 entries first, avoiding greedy mis-assignment for
# PLN/CZK pairs with nearly equal fuzzy errors.
matches = []
skips   = []

for p in sorted(all_payouts, key=lambda x: x["summe_eur"], reverse=True):
    summe   = p["summe_eur"]
    country = p["country"]
    valid   = COUNTRY_S1_LABELS[country]

    best, best_diff = None, float("inf")
    for e in s1_entries:
        if e["matched"]:
            continue
        if e["label"] not in valid:
            continue
        betrag = e["betrag"]
        if p["currency"] == "EUR":
            diff = abs(betrag - summe)
            if diff < 0.02 and diff < best_diff:
                best, best_diff = e, diff
        else:
            if betrag > 0:
                rel = abs(betrag - summe) / betrag
                if rel <= FUZZY_TOL and rel < best_diff:
                    best, best_diff = e, rel

    if best:
        best["matched"] = True
        matches.append((best, p))
        print(f"  MATCH  {p['country']:10s} {p['sheet']:26s}  "
              f"payout={summe:8.2f}EUR  →  Nr={best['nr']:>3}  ({best['label']} {best['betrag']:.2f})")
    else:
        skips.append(p)
        print(f"  SKIP   {p['country']:10s} {p['sheet']:26s}  "
              f"payout={summe:8.2f}EUR  (no Sheet1 match → next month or later)")

unmatched = [e for e in s1_entries if not e["matched"]]
if unmatched:
    print(f"\n⚠ Unmatched Sheet1 entries ({len(unmatched)}):")
    for e in unmatched:
        print(f"  Nr={e['nr']}  {e['label']}  {e['betrag']:.2f}")
else:
    print("\n✓ All Sheet1 Kaufland/KTO/IBAN entries matched.")

# ── Step 7: write Kaufland sheet ───────────────────────────────────────────
# Update header to 13 columns
HEADERS = [
    "Nr", "Marktplatz", "Netto_Sale", "Ust", "Brutto_Sale",
    "Erstattung(19%)", "Marktpaltzgebühr(19%)", "Commission(19%)",
    "Gutscheingebühr(19%)", "Steuer_Ohne_Gebühr(19%)", "Sonstiges(19%)",
    "KTO_Guthaben", "Behalten_in_KTO", "Summe",
]
for c, h in enumerate(HEADERS, 1):
    ws_kauf.cell(1, c).value = h

# Sheet1: ensure col 11 and 12 headers exist
if not ws_s1.cell(1, 11).value:
    ws_s1.cell(1, 11).value = "Steuersatz"
if not ws_s1.cell(1, 12).value:
    ws_s1.cell(1, 12).value = "Marktplatz"

# Clear old data (rows 2+)
for r in range(2, ws_kauf.max_row + 1):
    for c in range(1, 15):
        ws_kauf.cell(r, c).value = None

matches.sort(key=lambda x: x[0]["nr"])

print("\n=== Kaufland sheet rows written ===")
print(f"{'Nr':>4}  {'Marktplatz':<14}  {'Brutto':>10}  {'Summe':>10}  "
      f"{'S1_Betrag':>10}  {'S1_Diff':>8}  {'InnerChk':>9}")
print("─" * 82)

for i, (s1, p) in enumerate(matches):
    row = i + 2
    nr  = s1["nr"]
    mk  = COUNTRY_MARKT[p["country"]]

    tax_rate = p["tax_rate"]
    # Derive per-payout actual exchange rate from the bank Betrag.
    # For EUR rows actual_rate = 1.0 (rate_used); for CZK/PLN this eliminates
    # the monthly-average-vs-daily-rate discrepancy.
    if p["currency"] != "EUR" and p["summe_local"]:
        actual_rate = s1["betrag"] / p["summe_local"]
    else:
        actual_rate = p["rate_used"]

    if p["pure_balance"]:
        brutto = 0.0
        netto  = 0.0
        ust    = 0.0
        erstattung = marktplatz = commission = None
        gutschein  = steuer     = sonstiges  = None
    else:
        tl         = p["totals_local"]
        brutto     = r2(tl["brutto"]     * actual_rate)
        netto      = r2(brutto / (1 + tax_rate))
        ust        = r2(brutto - netto)
        erstattung = r2(tl["erstattung"] * actual_rate) or None
        marktplatz = r2(tl["marktplatz"] * actual_rate) or None
        commission = r2(tl["commission"] * actual_rate) or None
        gutschein  = r2(tl["gutschein"]  * actual_rate) or None
        steuer     = r2(tl["steuer"]     * actual_rate) or None
        sonstiges  = r2(tl["sonstiges"]  * actual_rate) or None
    kto_gut  = r2(p["kto_gut_local"]  * actual_rate) or None
    behalten = r2(p["behalten_local"] * actual_rate) or None
    summe    = s1["betrag"]

    ws_kauf.cell(row,  1).value = nr
    ws_kauf.cell(row,  2).value = mk
    ws_kauf.cell(row,  3).value = r2(netto)
    ws_kauf.cell(row,  4).value = r2(ust)
    ws_kauf.cell(row,  5).value = brutto
    ws_kauf.cell(row,  6).value = erstattung
    ws_kauf.cell(row,  7).value = marktplatz
    ws_kauf.cell(row,  8).value = commission
    ws_kauf.cell(row,  9).value = gutschein
    ws_kauf.cell(row, 10).value = steuer
    ws_kauf.cell(row, 11).value = sonstiges
    ws_kauf.cell(row, 12).value = kto_gut
    ws_kauf.cell(row, 13).value = behalten
    ws_kauf.cell(row, 14).value = summe

    # inner check: payout = brutto + fees + KTO_Guthaben − Behalten_in_KTO
    fields = [brutto,
              erstattung or 0, marktplatz or 0, commission or 0,
              gutschein  or 0, steuer     or 0, sonstiges  or 0,
              kto_gut    or 0, behalten   or 0]
    inner_diff = r2(r2(sum(fields)) - summe)

    # Sheet1 cols:
    #   5=Brrutto_Umsatz, 6=Andere Gebühren (19% fees only, excl. steuer),
    #   7=KTO_Guthaben, 8=Behalten, 10=Gebühren Ohne Steuer, 11=Steuersatz
    andere = r2(sum([erstattung or 0, marktplatz or 0, commission or 0,
                     gutschein  or 0, sonstiges  or 0]))  # steuer excluded
    ws_s1.cell(s1["row"], 5).value  = brutto or None
    ws_s1.cell(s1["row"], 6).value  = andere or None
    ws_s1.cell(s1["row"], 7).value  = kto_gut
    ws_s1.cell(s1["row"], 8).value  = behalten
    ws_s1.cell(s1["row"], 10).value = steuer
    ws_s1.cell(s1["row"], 11).value = tax_rate
    ws_s1.cell(s1["row"], 12).value = mk

    s1_diff = r2(s1["betrag"] - summe)
    rate_tag = f"  rate={actual_rate:.5f}" if p["currency"] != "EUR" else ""
    print(f"{nr:>4}  {mk:<14}  {brutto:>10.2f}  {summe:>10.2f}  "
          f"{s1['betrag']:>10.2f}  {s1_diff:>8.2f}  {inner_diff:>9.2f}{rate_tag}")

# ── Step 8: save ───────────────────────────────────────────────────────────
wb_us.save(UMSATZ_FILE)
print(f"\n✓ Saved: {UMSATZ_FILE}")
print(f"  {len(matches)} rows written to Kaufland sheet, {len(matches)} Sheet1 rows updated.")

# ── Step 9: write pending-payouts file ─────────────────────────────────────
monat_m      = re.search(r'Umsatzsteuer_(.+)\.xlsx', os.path.basename(UMSATZ_FILE))
monat        = monat_m.group(1) if monat_m else "unbekannt"
pending_file = os.path.join(os.path.dirname(UMSATZ_FILE) or ".", f"Kaufland_pending_{monat}.txt")
with open(pending_file, "w", encoding="utf-8") as pf:
    pf.write(f"# Kaufland 待到账 Payouts — {monat}\n")
    pf.write(f"# 以下 payout 在本月 Kontodruckansicht 中未出现，预计下月到账。\n")
    pf.write(f"# 执行下月做账时请核对 Sheet1 中是否已收到这些款项。\n\n")
    if skips:
        pf.write(f"{'国家/站点':<12}  {'来源Sheet':<32}  {'金额':>18}\n")
        pf.write("─" * 68 + "\n")
        for s in sorted(skips, key=lambda x: x["summe_eur"], reverse=True):
            if s["currency"] == "EUR":
                pf.write(f"{s['country']:<12}  {s['sheet']:<32}  {s['summe_eur']:>12.2f} EUR\n")
            else:
                pf.write(f"{s['country']:<12}  {s['sheet']:<32}  "
                         f"{s['summe_local']:>10.2f} {s['currency']}  ({s['summe_eur']:.2f} EUR)\n")
    else:
        pf.write("  (本月无待到账 payout)\n")
print(f"✓ Pending payouts saved: {pending_file}  ({len(skips)} entries)")
