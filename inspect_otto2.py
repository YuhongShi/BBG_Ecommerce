# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

KONTOAUSZUG = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Buchhaltung\Kontoauszug_Otto.xlsx"
UMSATZ = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_März.xlsx"

# Full Kontoauszug Otto sheet
wb_k = openpyxl.load_workbook(KONTOAUSZUG, data_only=True)
ws_k = wb_k["Otto"]
print("=== Kontoauszug Otto sheet (all rows) ===")
for i, row in enumerate(ws_k.iter_rows(min_row=1, values_only=True)):
    if any(v is not None for v in row):
        print(f"  row{i+1}: {row}")

# Sheet1 OTTO rows
wb_u = openpyxl.load_workbook(UMSATZ, data_only=True)
ws1 = wb_u["Sheet1"]
print("\n=== Sheet1 OTTO rows ===")
for i, row in enumerate(ws1.iter_rows(min_row=1, values_only=True)):
    if i == 0 or (row[2] == "OTTO"):
        print(f"  row{i+1}: {row}")
