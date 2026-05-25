# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

KONTOAUSZUG = r"D:\Workstation_Yuhong\Yuhong's_File\Python\Buchhaltung\Kontoauszug_Otto.xlsx"
UMSATZ = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_März.xlsx"

# ── Read Kontoauszug Otto sheet ────────────────────────────────────────────
wb_k = openpyxl.load_workbook(KONTOAUSZUG, data_only=True)
ws_k = wb_k["Otto"]

konto_rows = {}  # Nr -> row tuple (0-indexed)
for row in ws_k.iter_rows(min_row=2, values_only=True):
    nr = row[0]
    if nr is not None:
        konto_rows[nr] = row

# ── Open Umsatzsteuer workbook ─────────────────────────────────────────────
wb_u = openpyxl.load_workbook(UMSATZ)
ws_otto = wb_u["OTTO"]
ws1 = wb_u["Sheet1"]

# ── Patch: set VAT-Korrektur for Nr=9, recalculate Summe ──────────────────
VAT_PATCH = {9: -11.22}

for row in ws_otto.iter_rows(min_row=2):
    nr = row[0].value
    if nr in VAT_PATCH:
        vat = VAT_PATCH[nr]
        row[10].value = vat  # col 11 = VAT-Korrektur
        # Summe = sum of cols 5-10 (Brutto_Sale through VAT-Korrektur)
        # = Kontoauszug Summe + VAT-Korrektur
        old_summe = konto_rows[nr][10]
        new_summe = round(old_summe + vat, 2)
        row[11].value = new_summe  # col 12 = Summe
        print(f"Patched Nr={nr}: VAT-Korrektur={vat}, Summe {old_summe} -> {new_summe}")

# ── Step 2: Verify ────────────────────────────────────────────────────────
# Rebuild OTTO summe map from sheet (after patch)
otto_summe = {}
for row in ws_otto.iter_rows(min_row=2, values_only=True):
    nr = row[0]
    if nr is not None:
        otto_summe[nr] = row[11]  # col 12

sheet1_otto = {}
for row in ws1.iter_rows(min_row=2, values_only=True):
    nr, zahlungs, betrag = row[0], row[2], row[3]
    if zahlungs == "OTTO" and nr is not None:
        sheet1_otto[nr] = betrag

print("\n=== Step 2: Verification ===")
mismatches = []
for nr, betrag in sheet1_otto.items():
    summe = otto_summe.get(nr)
    b_r = round(float(betrag), 2) if betrag is not None else None
    s_r = round(float(summe), 2) if summe is not None else None
    ok = b_r == s_r
    mark = "[OK]" if ok else "[NG]"
    print(f"  {mark} Nr={nr}  Sheet1 Betrag={betrag}  OTTO Summe={summe}")
    if not ok:
        mismatches.append((nr, betrag, summe))

if mismatches:
    print("\n[!] Still mismatched — not proceeding.")
    for nr, b, s in mismatches:
        print(f"  Nr={nr}  Betrag={b}  Summe={s}  diff={round(float(s)-float(b),2)}")
    wb_u.save(UMSATZ)
    sys.exit(1)

print("All rows match [OK]")

# ── Step 3: Fill Sheet1 Brrutto_Umsatz (col5) and Andere Gebuehren (col6) ─
# OTTO sheet (0-indexed): col4=Brutto_Sale, col5-10 = fees incl VAT-Korrektur
otto_calc = {}
for row in ws_otto.iter_rows(min_row=2, values_only=True):
    nr = row[0]
    if nr is None:
        continue
    brutto = row[4] or 0
    andere = sum(row[i] or 0 for i in range(5, 11))
    otto_calc[nr] = (brutto, round(andere, 2))

print(f"\n=== Step 3: Brrutto_Umsatz & Andere Gebuehren ===")
print(f"  {'Nr':<6} {'Brrutto_Umsatz':>16} {'Andere Gebuehren':>18}")
print("  " + "-" * 42)

for row in ws1.iter_rows(min_row=2):
    nr = row[0].value
    zahlungs = row[2].value
    if zahlungs == "OTTO" and nr is not None and nr in otto_calc:
        brutto, andere = otto_calc[nr]
        row[4].value = brutto
        row[5].value = andere
        print(f"  Nr={str(nr):<4} {str(brutto):>16} {str(andere):>18}")

wb_u.save(UMSATZ)
print(f"\nSaved: {UMSATZ}")
print("Done.")
