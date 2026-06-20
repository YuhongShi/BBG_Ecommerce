#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Show ALL rows (including nulls) around payout groups for problem countries."""
import sys, re
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl

def parse_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try: return float(s)
    except: return None

KAUFLAND_FILE = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Kaufland\2026年\5月\kaufland-BBG 26年5月回款分佣及vat应结算金额-41772.49欧元(施总)(已发送确认中).xlsx"
wb = openpyxl.load_workbook(KAUFLAND_FILE, data_only=True)

# Check these countries
TARGET_COUNTRIES = ["奥地利站", "意大利站", "斯洛伐克站", "法国站"]
TARGET_SUMME = {
    "奥地利站":   689.51,
    "意大利站":   131.14,
    "斯洛伐克站": 237.33,
    "法国站":      36.68,
}

for sn in wb.sheetnames:
    country = next((k for k in TARGET_COUNTRIES if k in sn), None)
    if not country:
        continue
    # Only look at May sheets
    m = re.search(r'(\d+)年(\d+)月', sn)
    if not m or int(m.group(2)) != 5:
        continue

    ws = wb[sn]
    # Find header row
    hdr = next(
        (r for r in range(1, 12) if str(ws.cell(r, 1).value or "").strip() == "booking_date"),
        None
    )
    if hdr is None:
        print(f"\n{sn}: header not found")
        continue

    # Read ALL rows including nulls
    all_rows = []
    for r in range(hdr + 1, ws.max_row + 1):
        bt     = ws.cell(r, 2).value
        date   = ws.cell(r, 1).value
        amount = parse_num(ws.cell(r, 5).value)
        bal    = parse_num(ws.cell(r, 6).value)
        if date is None and bt is None and amount is None and bal is None:
            break
        all_rows.append({"row": r, "date": date, "bt": bt, "amount": amount, "balance": bal})

    target_summe = TARGET_SUMME[country]

    # Find the payout row near target_summe
    print(f"\n{'='*80}")
    print(f"Sheet: {sn}  (target payout ≈ {target_summe:.2f} EUR)")
    print(f"Total rows: {len(all_rows)}")
    print()

    # Find payout rows
    payout_rows = [i for i, r in enumerate(all_rows) if str(r["bt"] or "").strip() == "Payout"]
    if not payout_rows:
        print("  No Payout rows found.")
        continue

    # Find the payout closest to target_summe
    target_idx = None
    for pi in payout_rows:
        amt = all_rows[pi]["amount"]
        if amt is not None and abs(abs(amt) - target_summe) < 1.0:
            target_idx = pi
            break

    if target_idx is None:
        print(f"  Could not find payout matching {target_summe:.2f}")
        print(f"  Payout amounts: {[abs(all_rows[i]['amount']) for i in payout_rows if all_rows[i]['amount']]}")
        continue

    # Find start of this group (previous payout or beginning)
    prev_payout_idx = max([i for i in payout_rows if i < target_idx], default=None)
    start = (prev_payout_idx + 1) if prev_payout_idx is not None else 0

    print(f"  Printing rows {start} to {target_idx} (group for payout ≈ {target_summe:.2f}):")
    print(f"  {'#':>4}  {'date':<12}  {'booking_text':<55}  {'amount':>12}  {'balance':>12}")
    print(f"  {'-'*110}")

    running_sum = 0.0
    null_bt_sum = 0.0
    for i in range(start, target_idx + 1):
        r = all_rows[i]
        bt_str   = str(r["bt"] or "")[:55]
        amt      = r["amount"] or 0.0
        bal      = r["balance"]
        marker   = " ← NULL bt" if r["bt"] is None and amt != 0 else ""
        null_indicator = "***" if r["bt"] is None and amt != 0 else "   "
        if r["bt"] is not None:
            running_sum += amt
        else:
            null_bt_sum += amt if amt else 0
        bal_str = f"{bal:.4f}" if bal is not None else "None"
        print(f"  {null_indicator}{i+1:>3}  {str(r['date'] or ''):<12}  {bt_str:<55}  {amt:>12.4f}  {bal_str:>12}")

    print(f"\n  Sum of rows WITH booking_text: {running_sum:.4f}")
    print(f"  Sum of rows WITH NULL bt:      {null_bt_sum:.4f}")
    print(f"  Payout amount:                 {abs(all_rows[target_idx]['amount']):.4f}")
    print(f"  Balance after payout:          {all_rows[target_idx]['balance']:.4f}")
