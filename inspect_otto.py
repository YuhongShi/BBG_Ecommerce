# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

UMSATZ = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_März.xlsx"

wb = openpyxl.load_workbook(UMSATZ, data_only=True)
print("Sheets:", wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== Sheet: {sname} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
        print(f"  row{i+1}: {row[:14]}")
