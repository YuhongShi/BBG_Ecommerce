# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

KONTOAUSZUG = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Buchhaltung\Kontoauszug_Ebay.xlsx"
UMSATZ = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_März.xlsx"

# ── Read Kontoauszug Ebay_Umsatz ───────────────────────────────────────────
# Cols: Nr(1), Marktplatz(2), Netto_Sale(3), Ust(4), Brutto_Sale(5),
#       Erstattung(6), Andere Gebuehre(7), Einbehalten(8), Commission(9),
#       Internationale Gebuehr(10), Einzahlung(11), Summe(12)
wb_k = openpyxl.load_workbook(KONTOAUSZUG, data_only=True)
ws_k = wb_k["Ebay_Umsatz"]

konto_rows = {}  # Nr -> row tuple (0-indexed)
for row in ws_k.iter_rows(min_row=2, values_only=True):
    nr = row[0]
    if nr is not None:
        konto_rows[nr] = row

print(f"Kontoauszug: {len(konto_rows)} data rows")

# ── Open Umsatzsteuer workbook ─────────────────────────────────────────────
wb_u = openpyxl.load_workbook(UMSATZ)
ws_ebay = wb_u["Ebay"]
ws1 = wb_u["Sheet1"]

# ── Step 1: Write Kontoauszug data into Ebay sheet ────────────────────────
# Umsatzsteuer Ebay cols: same as Kontoauszug cols 1-12 (Summe at col 12)
# (col 13=None, col 14=check — leave untouched)

# Build existing Nr set in Ebay sheet
existing_nr_rows = {}  # Nr -> row_number
for row in ws_ebay.iter_rows(min_row=2):
    nr = row[0].value
    if nr is not None:
        existing_nr_rows[nr] = row[0].row

# Find next empty row
next_row = 2
for row in ws_ebay.iter_rows(min_row=2):
    if row[0].value is not None:
        next_row = row[0].row + 1
    else:
        break

written = []
for nr, r in konto_rows.items():
    if nr in existing_nr_rows:
        row_num = existing_nr_rows[nr]
        for col_i in range(12):
            ws_ebay.cell(row=row_num, column=col_i+1, value=r[col_i])
        written.append((nr, "updated"))
    else:
        for col_i in range(12):
            ws_ebay.cell(row=next_row, column=col_i+1, value=r[col_i])
        written.append((nr, "inserted"))
        next_row += 1

print(f"\n=== Step 1: {len(written)} rows written to Ebay sheet ===")
for nr, action in written:
    summe = konto_rows[nr][11]
    print(f"  Nr={nr} [{action}]  Summe={summe}")

# ── Step 2: Verify vs Sheet1 Betrag ───────────────────────────────────────
# Rebuild Ebay summe map from updated sheet
ebay_summe = {}
for row in ws_ebay.iter_rows(min_row=2, values_only=True):
    nr = row[0]
    if nr is not None:
        ebay_summe[nr] = row[11]  # col 12 = Summe

# Sheet1 eBay rows (col 3 = 'eBay')
sheet1_ebay = {}
for row in ws1.iter_rows(min_row=2, values_only=True):
    nr, zahlungs, betrag = row[0], row[2], row[3]
    if zahlungs == "eBay" and nr is not None:
        sheet1_ebay[nr] = betrag

print(f"\n=== Step 2: Verification ({len(sheet1_ebay)} Sheet1 eBay rows) ===")
mismatches = []
missing_in_konto = []

for nr, betrag in sheet1_ebay.items():
    if nr not in konto_rows:
        missing_in_konto.append((nr, betrag))
        print(f"  [??] Nr={nr}  Sheet1 Betrag={betrag}  -- NOT IN Kontoauszug")
        continue
    summe = ebay_summe.get(nr)
    b_r = round(float(betrag), 2) if betrag is not None else None
    s_r = round(float(summe), 2) if summe is not None else None
    ok = b_r == s_r
    mark = "[OK]" if ok else "[NG]"
    diff = round(float(betrag) - float(summe), 2) if summe is not None else "N/A"
    print(f"  {mark} Nr={nr}  Sheet1 Betrag={betrag}  Ebay Summe={summe}  diff={diff}")
    if not ok:
        mismatches.append((nr, betrag, summe, diff))

if mismatches or missing_in_konto:
    print("\n[!] Issues found — stopping. Please advise:")
    if missing_in_konto:
        print("\n  Missing in Kontoauszug (no data to fill):")
        for nr, b in missing_in_konto:
            print(f"    Nr={nr}  Sheet1 Betrag={b}")
    if mismatches:
        print(f"\n  {'Nr':<8} {'Sheet1 Betrag':>15} {'Ebay Summe':>12} {'diff':>8}")
        print("  " + "-" * 46)
        for nr, b, s, d in mismatches:
            print(f"  {str(nr):<8} {str(b):>15} {str(s):>12} {str(d):>8}")
    wb_u.save(UMSATZ)
    sys.exit(1)

print("\nAll rows match [OK] — proceeding to Step 3")

# ── Step 3: Fill Sheet1 Brrutto_Umsatz (col5) and Andere Gebuehren (col6) ─
# Ebay sheet (0-indexed): col4=Brutto_Sale, cols 5-10 = fees
ebay_calc = {}
for row in ws_ebay.iter_rows(min_row=2, values_only=True):
    nr = row[0]
    if nr is None:
        continue
    brutto = row[4] or 0
    andere = sum(row[i] or 0 for i in range(5, 11))
    ebay_calc[nr] = (brutto, round(andere, 2))

print(f"\n=== Step 3: Brrutto_Umsatz & Andere Gebuehren ===")
print(f"  {'Nr':<6} {'Brrutto_Umsatz':>16} {'Andere Gebuehren':>18}")
print("  " + "-" * 43)

for row in ws1.iter_rows(min_row=2):
    nr = row[0].value
    zahlungs = row[2].value
    if zahlungs == "eBay" and nr is not None and nr in ebay_calc:
        brutto, andere = ebay_calc[nr]
        row[4].value = brutto
        row[5].value = andere
        print(f"  Nr={str(nr):<4} {str(brutto):>16} {str(andere):>18}")

wb_u.save(UMSATZ)
print(f"\nSaved: {UMSATZ}")
print("Done.")
