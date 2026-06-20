"""
从 Kontodruckansicht_Mai.pdf 手工提取的交易数据写入 Umsatzsteuer_Mai.xlsx Sheet1
Col 1=Nr, Col 2=Datum (YYYY.MM.DD), Col 3=Zahlungsverkehr, Col 4=Betrag
"""
import shutil
from openpyxl import load_workbook

SRC = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_April.xlsx"
TARGET = r"D:\Workstation_Yuhong\BBG\电商\网买做账\Umsatzsteuer_Mai.xlsx"

TRANSACTIONS = [
    (1,  "2026.05.26", "eBay",              173.22),
    (2,  "2026.05.26", "Amazon",             91.18),
    (3,  "2026.05.26", "XPOLLENS",         5062.73),
    (4,  "2026.05.26", "Orber Germany GmbH", -5746.24),
    (5,  "2026.05.26", "Orber Germany GmbH", -8209.02),
    (6,  "2026.05.26", "Orber Germany GmbH", -16924.60),
    (7,  "2026.05.26", "Orber Germany GmbH", -6176.47),
    (8,  "2026.05.26", "Orber Germany GmbH",  -126.06),
    (9,  "2026.05.26", "TikTok",             158.08),
    (10, "2026.05.26", "IBAN",               199.42),
    (11, "2026.05.26", "Kaufland",           250.13),
    (12, "2026.05.26", "Kaufland",          5478.19),
    (13, "2026.05.26", "eBay",               531.34),
    (14, "2026.05.26", "eBay",               450.63),
    (15, "2026.05.26", "eBay",               190.53),
    (16, "2026.05.22", "eBay",               428.72),
    (17, "2026.05.22", "KTO",                214.91),
    (18, "2026.05.22", "BBG Germany GmbH",  -5000.00),
    (19, "2026.05.21", "OTTO",              4353.80),
    (20, "2026.05.21", "eBay",               243.51),
    (21, "2026.05.20", "Kaufland",           779.14),
    (22, "2026.05.20", "eBay",               698.84),
    (23, "2026.05.20", "KTO",                183.27),
    (24, "2026.05.19", "Kaufland",            87.79),
    (25, "2026.05.19", "Kaufland",            67.82),
    (26, "2026.05.19", "Kaufland",            36.68),
    (27, "2026.05.19", "Kaufland",          1417.17),
    (28, "2026.05.19", "eBay",               606.75),
    (29, "2026.05.19", "KTO",                276.47),
    (30, "2026.05.19", "IBAN",               386.05),
    (31, "2026.05.18", "XPOLLENS",         18525.80),
    (32, "2026.05.18", "Kaufland",          1000.00),
    (33, "2026.05.18", "Kaufland",          2348.88),
    (34, "2026.05.18", "Kaufland",           232.92),
    (35, "2026.05.18", "Kaufland",           317.79),
    (36, "2026.05.18", "eBay",               441.28),
    (37, "2026.05.18", "eBay",               387.91),
    (38, "2026.05.18", "eBay",               450.54),
    (39, "2026.05.15", "OTTO",              4341.44),
    (40, "2026.05.15", "eBay",               164.66),
    (41, "2026.05.15", "eBay",               336.21),
    (42, "2026.05.13", "Kaufland",          1919.94),
    (43, "2026.05.13", "Kaufland",           168.98),
    (44, "2026.05.13", "eBay",               658.13),
    (45, "2026.05.13", "KTO",                378.66),
    (46, "2026.05.13", "KTO",                790.73),
    (47, "2026.05.13", "IBAN",                51.41),
    (48, "2026.05.12", "Kaufland",           119.59),
    (49, "2026.05.12", "eBay",               390.60),
    (50, "2026.05.12", "IBAN",              1134.36),
    (51, "2026.05.11", "XPOLLENS",         11267.78),
    (52, "2026.05.11", "Kaufland",          1505.69),
    (53, "2026.05.11", "Kaufland",           163.12),
    (54, "2026.05.11", "eBay",                67.79),
    (55, "2026.05.11", "eBay",               369.45),
    (56, "2026.05.11", "eBay",               200.85),
    (57, "2026.05.08", "Kaufland",           131.14),
    (58, "2026.05.08", "Kaufland",           237.33),
    (59, "2026.05.08", "Kaufland",          6411.69),
    (60, "2026.05.08", "Kaufland",           689.51),
    (61, "2026.05.08", "eBay",               783.53),
    (62, "2026.05.07", "OTTO",              3540.06),
    (63, "2026.05.07", "eBay",               202.86),
    (64, "2026.05.06", "XPOLLENS",          2457.24),
    (65, "2026.05.06", "eBay",               206.91),
    (66, "2026.05.06", "KTO",                192.85),
    (67, "2026.05.06", "IBAN",               653.00),
    (68, "2026.05.05", "eBay",               328.88),
    (69, "2026.05.05", "IBAN",               279.10),
    (70, "2026.05.04", "Kaufland",           174.13),
    (71, "2026.05.04", "Kaufland",           218.51),
    (72, "2026.05.04", "Kaufland",          2093.76),
    (73, "2026.05.04", "Kaufland",           138.78),
    (74, "2026.05.04", "eBay",              1189.13),
    (75, "2026.05.04", "eBay",               186.49),
    (76, "2026.05.04", "eBay",               153.44),
]

# Copy April → Mai
shutil.copy2(SRC, TARGET)
print(f"Copied April -> Mai")

wb = load_workbook(TARGET)
ws = wb["Sheet1"]

# Clear cols A-D from row 2 onwards
for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
    for cell in row:
        cell.value = None

# Write headers in row 1 only if the cells are empty
if not ws.cell(1, 1).value:
    ws.cell(1, 1).value = "Nr"
if not ws.cell(1, 2).value:
    ws.cell(1, 2).value = "Datum"
if not ws.cell(1, 3).value:
    ws.cell(1, 3).value = "Zahlungsverkehr"
if not ws.cell(1, 4).value:
    ws.cell(1, 4).value = "Betrag"

for nr, date, zv, betrag in TRANSACTIONS:
    row = nr + 1
    ws.cell(row, 1).value = nr
    ws.cell(row, 2).value = date
    ws.cell(row, 3).value = zv
    ws.cell(row, 4).value = betrag

wb.save(TARGET)
print(f"Done: wrote {len(TRANSACTIONS)} rows (Nr 1-{len(TRANSACTIONS)})")
