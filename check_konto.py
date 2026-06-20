import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r"D:\Workstation_Yuhong\Yuhong's_File\Python\Buchhaltung\Kontoauszug_Otto.xlsx", data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb['Otto']
print('Headers:', [c.value for c in ws[1]])
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    print(row)
